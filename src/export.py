from __future__ import annotations

import sqlite3
import io
from dataclasses import dataclass
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from src.config import DB_PATH, WEEKDAYS, SLOTS
from src.utils import gc
WD_SHORT = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]

STATE_NORMAL = "normal"
STATE_CANCELLED = "cancelled"
STATE_TRANSFERRED = "transferred"
STATE_RESTORED = "restored"


@dataclass
class CellEntry:
    text: str
    state: str



def _slot_by_start(start: str) -> int:
    s = start[11:16] if len(start) > 5 else start
    for i, sl in enumerate(SLOTS):
        if sl["start"] == s:
            return i
    return -1


def get_all_groups() -> list[str]:
    conn = gc()
    rows = conn.execute("SELECT name FROM groups ORDER BY name").fetchall()
    conn.close()
    return [r["name"] for r in rows]


def _get_cancellation_info(conn, sids: list[int]) -> dict:
    if not sids:
        return {"cancelled_sids": set(), "cancelled_restored_sids": set(), "restored_schedule_ids": set()}

    ph = ",".join("?" for _ in sids)
    rows = conn.execute(f"""
        SELECT schedule_id, is_restored, restored_schedule_id
        FROM cancellations
        WHERE schedule_id IN ({ph})
    """, sids).fetchall()

    cancelled_sids = set()
    cancelled_restored_sids = set()
    restored_schedule_ids = set()

    for r in rows:
        if r["is_restored"]:
            cancelled_restored_sids.add(r["schedule_id"])
            if r["restored_schedule_id"]:
                restored_schedule_ids.add(r["restored_schedule_id"])
        else:
            cancelled_sids.add(r["schedule_id"])

    return {
        "cancelled_sids": cancelled_sids,
        "cancelled_restored_sids": cancelled_restored_sids,
        "restored_schedule_ids": restored_schedule_ids,
    }


def _get_transfers_map(conn, sids: list[int]) -> dict[int, dict]:
    if not sids:
        return {}
    ph = ",".join("?" for _ in sids)
    rows = conn.execute(f"""
        SELECT t.schedule_id, r.name AS new_room_name, r.building AS new_room_building
        FROM transfers t
        JOIN rooms r ON t.new_room_id = r.id
        WHERE t.schedule_id IN ({ph})
    """, sids).fetchall()
    return {r["schedule_id"]: dict(r) for r in rows}


def _build_schedule_grid(
    entries: list[dict],
    transfers_map: dict[int, dict],
    cancelled_sids: set[int],
    cancelled_restored_sids: set[int],
    restored_schedule_ids: set[int],
    entity_type: str,
) -> dict[str, dict[str, dict[int, list[CellEntry]]]]:
    grid: dict[str, dict[str, dict[int, list[CellEntry]]]] = {}
    for wt_key in ("upper", "lower"):
        wt_label = "Верхняя неделя" if wt_key == "upper" else "Нижняя неделя"
        grid[wt_label] = {}
        for wd in WEEKDAYS:
            grid[wt_label][wd] = defaultdict(list)

    for e in entries:
        sid = e["schedule_id"]

        wt_key = e["week_type"]
        wt_label = "Верхняя неделя" if wt_key == "upper" else "Нижняя неделя"
        wd = e["weekday"]
        slot_idx = _slot_by_start(e["start"])
        if slot_idx < 0:
            continue

        title = e["lesson_title"]
        ltype = e["lesson_type"]
        transfer = transfers_map.get(sid)

        if entity_type == "group":
            entity_info = e.get("teacher", "") or ""
        else:
            entity_info = e.get("group_name", "") or ""

        if sid in cancelled_sids:
            state = STATE_CANCELLED
            room_str = f"{e['room_name']} (корп. {e['room_building']})"
            parts = [f"{title} ({ltype})"]
            if entity_info:
                parts.append(entity_info)
            parts.append(room_str)
            parts.append("ОТМЕНЕНО")
        elif sid in cancelled_restored_sids:
            state = STATE_CANCELLED
            room_str = f"{e['room_name']} (корп. {e['room_building']})"
            parts = [f"{title} ({ltype})"]
            if entity_info:
                parts.append(entity_info)
            parts.append(room_str)
            parts.append("ОТМЕНЕНО → ВОССТАНОВЛЕНО")
        elif sid in restored_schedule_ids:
            state = STATE_RESTORED
            room_str = f"{e['room_name']} (корп. {e['room_building']})"
            parts = [f"{title} ({ltype})"]
            if entity_info:
                parts.append(entity_info)
            parts.append(room_str)
            parts.append("ВОССТАНОВЛЕНО")
        elif transfer:
            state = STATE_TRANSFERRED
            room_str = f"{transfer['new_room_name']} (корп. {transfer['new_room_building']})"
            parts = [f"{title} ({ltype})"]
            if entity_info:
                parts.append(entity_info)
            parts.append(f"→ {room_str} (перенос из {e['room_name']})")
        else:
            state = STATE_NORMAL
            room_str = f"{e['room_name']} (корп. {e['room_building']})"
            parts = [f"{title} ({ltype})"]
            if entity_info:
                parts.append(entity_info)
            parts.append(room_str)

        cell_text = "\n".join(parts)
        grid[wt_label][wd][slot_idx].append(CellEntry(text=cell_text, state=state))

    return grid


def get_schedule_for_group(group_name: str) -> dict[str, dict[str, dict[int, list[CellEntry]]]]:
    conn = gc()

    entries = conn.execute("""
        SELECT s.id AS schedule_id, s.weekday, s.start, s.end, s.week_type,
               l.title AS lesson_title, l.lesson_type, l.teacher,
               r.name AS room_name, r.building AS room_building
        FROM schedule s
        JOIN lessons l ON s.lesson_id = l.id
        JOIN groups g ON s.group_id = g.id
        JOIN rooms r ON s.room_id = r.id
        WHERE g.name = ?
        ORDER BY s.week_type, s.weekday, s.start
    """, (group_name,)).fetchall()

    sids = [r["schedule_id"] for r in entries]

    canc_info = _get_cancellation_info(conn, sids)
    transfers_map = _get_transfers_map(conn, sids)

    conn.close()

    return _build_schedule_grid(
        [dict(r) for r in entries],
        transfers_map,
        canc_info["cancelled_sids"],
        canc_info["cancelled_restored_sids"],
        canc_info["restored_schedule_ids"],
        entity_type="group",
    )


def get_schedule_for_teacher(teacher: str) -> dict[str, dict[str, dict[int, list[CellEntry]]]]:
    conn = gc()

    entries = conn.execute("""
        SELECT s.id AS schedule_id, s.weekday, s.start, s.end, s.week_type,
               l.title AS lesson_title, l.lesson_type, l.teacher,
               g.name AS group_name,
               r.name AS room_name, r.building AS room_building
        FROM schedule s
        JOIN lessons l ON s.lesson_id = l.id
        JOIN groups g ON s.group_id = g.id
        JOIN rooms r ON s.room_id = r.id
        WHERE l.teacher LIKE ?
        ORDER BY s.week_type, s.weekday, s.start
    """, (f"%{teacher}%",)).fetchall()

    sids = [r["schedule_id"] for r in entries]

    canc_info = _get_cancellation_info(conn, sids)
    transfers_map = _get_transfers_map(conn, sids)

    conn.close()

    return _build_schedule_grid(
        [dict(r) for r in entries],
        transfers_map,
        canc_info["cancelled_sids"],
        canc_info["cancelled_restored_sids"],
        canc_info["restored_schedule_ids"],
        entity_type="teacher",
    )


def generate_excel(
    grid: dict[str, dict[str, dict[int, list[CellEntry]]]],
    title: str,
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    slot_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    header_font = Font(bold=True, size=11)

    normal_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    normal_font = Font(size=10)
    transferred_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    transferred_font = Font(color="065F46", size=10)
    cancelled_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    cancelled_font = Font(color="999999", strikethrough=True, size=10)
    restored_fill = PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid")
    restored_font = Font(color="5B21B6", size=10)

    wrap_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    state_styles = {
        STATE_NORMAL: (normal_fill, normal_font),
        STATE_CANCELLED: (cancelled_fill, cancelled_font),
        STATE_TRANSFERRED: (transferred_fill, transferred_font),
        STATE_RESTORED: (restored_fill, restored_font),
    }
    state_priority = {STATE_CANCELLED: 0, STATE_RESTORED: 1, STATE_TRANSFERRED: 2, STATE_NORMAL: 3}

    for sheet_name in ("Верхняя неделя", "Нижняя неделя"):
        ws = wb.create_sheet(title=sheet_name)

        ws.cell(row=1, column=1, value="Пара / Время").font = header_font_white
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).border = thin_border
        ws.cell(row=1, column=1).alignment = center_align

        for col_idx, wd_short in enumerate(WD_SHORT, start=2):
            cell = ws.cell(row=1, column=col_idx, value=wd_short)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        for row_idx, slot in enumerate(SLOTS, start=2):
            label = f"{slot['name']}\n{slot['start']}–{slot['end']}"
            cell = ws.cell(row=row_idx, column=1, value=label)
            cell.font = header_font
            cell.fill = slot_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            max_lines = 1
            for col_idx, wd in enumerate(WEEKDAYS, start=2):
                cell_entries: list[CellEntry] = grid.get(sheet_name, {}).get(wd, {}).get(row_idx - 2, [])

                c = ws.cell(row=row_idx, column=col_idx)
                c.border = thin_border

                if cell_entries:
                    dominant = min(cell_entries, key=lambda e: state_priority[e.state]).state
                    fill, font = state_styles[dominant]

                    cell_text = "\n---\n".join(e.text for e in cell_entries)
                    c.value = cell_text
                    c.fill = fill
                    c.font = font
                    c.alignment = wrap_align

                    col_chars = 30
                    text_lines = 0
                    for line in cell_text.split("\n"):
                        line_len = sum(2 if ord(ch) > 127 else 1 for ch in line)
                        text_lines += max(1, -(-line_len // col_chars))
                    max_lines = max(max_lines, text_lines)

            ws.row_dimensions[row_idx].height = max(15 * max_lines, 30)

        ws.column_dimensions["A"].width = 14
        for col_idx in range(2, 9):
            ws.column_dimensions[get_column_letter(col_idx)].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
